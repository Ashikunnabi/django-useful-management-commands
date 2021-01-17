import json
import os
from collections import Counter
from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import make_password
from django.core.management.base import BaseCommand
from openpyxl import load_workbook  # pip install openpyxl

from your_app.models import Role


class Command(BaseCommand):
    help = 'Create/Add bulk user from excel'
    User = get_user_model()

    def add_arguments(self, parser):
        """ Command line arguments """
        parser.add_argument('-l', '--location', type=str)

    def get_excel_file_location(self, excel_location):
        """ Validate given excel location is correct """
        if excel_location:
            self.stdout.write('[+] Looking for excel file...')

            if not os.path.exists(excel_location):
                self.stderr.write("[X] Excel file location doesn't exists.")
            else:
                # check file extension for excel file
                if not excel_location.endswith('.xlsx'):
                    self.stderr.write("[X] This is not an excel file.")
                else:
                    self.stdout.write(
                        self.style.SUCCESS("[+] Excel file found successfully.")
                    )
                    return excel_location
        return ''

    def check_empty_value_in_column(self, column_header, row_data):
        """ Check excel has empty value in selected column """
        _data = [d[0].value for d in row_data]
        if '' in _data or None in _data:
            self.stderr.write(f'[X] "{column_header}" has empty value.')
            return False
        return True

    def check_duplicate_value_in_column(self, column_header, row_data):
        """ Check excel has duplicate value in selected column """
        _data = [d[0].value for d in row_data]
        if len(_data) != len(set(_data)):
            self.stderr.write(
                f'[X] "{column_header}" has duplicate value: \n'
                f'{[item+" "+ str(count) + " times" for item, count in Counter(_data).items() if count > 1]}'
            )
            return False
        return True

    def validate(self, columns, data):
        """ Validate excel files """
        for column in columns:
            if column[3]:
                rows = data.iter_rows(min_row=2, min_col=column[1], max_col=column[1])
                if not self.check_empty_value_in_column(column[2], rows):
                    return False
            if column[4]:
                rows = data.iter_rows(min_row=2, min_col=column[1], max_col=column[1])
                if not self.check_duplicate_value_in_column(column[2], rows):
                    return False
        return True

    def validate_excel_file(self, data):
        """ Prepare data for validate """
        self.stdout.write('[+] Validating excel file...')

        # prepare columns information for validation
        columns = [
            # (min_row, column, column_header, check_empty, check_duplicate)
            (2, 1, 'First Name', True, False),
            (2, 2, 'Last Name', True, False),
            (2, 3, 'Username', True, True),
            (2, 4, 'Email', True, True),
            (2, 6, 'Phone', True, False),
            (2, 8, 'Status', True, False),
            (2, 9, 'Role', True, False),
            (2, 10, 'Password', True, False)
        ]
        if self.validate(columns, data) is False:
            return False

        self.stdout.write(
            self.style.SUCCESS("[+] Excel file validated successfully.")
        )
        return True

    def user_create(self, data):
        """ Create users """
        user_list, error_details = [], []
        self.stdout.write('[+] User creating...')
        # take from 2nd row as 1st row contains headers
        rows = data.iter_rows(min_row=2)
        for row in rows:
            try:
                user = self.User(
                    first_name=row[0].value,
                    last_name=row[1].value,
                    username=row[2].value,
                    email=row[3].value,
                    address=row[4].value,
                    phone_number=row[5].value,
                    position=row[6].value,
                    status=row[7].value,
                    role=Role.objects.get(code=row[8].value),
                    password=make_password(row[9].value)
                )
                # validate user data
                user.full_clean()
                user_list.append(user)
            except Exception as ex:
                error = {
                    f'Row {row[0].row}': ex.__str__()
                }
                error_details.append(error)

        # Validation error found while model validation checking
        if error_details:
            self.stderr.write("[+] Validation error found. Please fix bellow issues:")
            self.stdout.write(json.dumps(error_details, indent=4, sort_keys=True))
            return
        self.User.objects.bulk_create(user_list)

        self.stdout.write(
            self.style.SUCCESS(f"[+] Total {len(user_list)} users created successfully.")
        )

    def handle(self, *args, **options):
        excel_location = options.get('location')

        # get proper file location
        excel_file_location = self.get_excel_file_location(excel_location)

        # if files location is not valid show error messages
        if excel_file_location == '':
            self.stderr.write(
                "[X] Please provide correct locations."
            )
            return

        # load excel file
        excel_file = load_workbook(excel_file_location)
        # read firs sheet
        data = excel_file.active

        # Validate excel data
        is_excel_valid = self.validate_excel_file(data)
        if is_excel_valid is False:
            return

        # Create new user
        self.user_create(data)

